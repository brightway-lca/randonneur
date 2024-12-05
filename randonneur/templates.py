from collections import defaultdict
from pathlib import Path
from typing import List, Optional

import openpyxl
import xlsxwriter

from randonneur import Datapackage, MappingConstants
from randonneur.licenses import LICENSES as ALL_LICENSES

ROLES = [
    "author",
    "publisher",
    "maintainer",
    "wrangler",
    "contributor",
]
MAPPINGS = [
    "SIMAPRO_CSV",
    "ECOSPOLD2",
    "ECOSPOLD1_BIO",
    "ECOSPOLD2_BIO",
    "ECOSPOLD2_BIO_FLOWMAPPER",
    "CUSTOM",
]
LICENSES = [
    "CC-BY-4.0",
    "CC-BY-NC-SA-4.0",
    "CC-BY-ND-4.0",
    "CC-BY-SA-4.0",
    "CC0-1.0",
    "CDL-1.0",
    "MIT",
    "ODC-By-1.0",
    "ODbL-1.0",
    "OML",
    "OSL-3.0",
    "PDDL-1.0",
    "PROPRIETARY",
]


def create_excel_template(data: List[dict], filepath: Path, replace_existing: bool = False) -> Path:
    """
    Create an Excel template with optionally some data for a new matching data file.

    `data` should be a list of dictionaries like `{'source': {}, 'target': {}}`. The keys and values
    in these sub-dictionaries should be strings, or castable to strings in a way that can be
    reversed. This function doesn't do any type conversion or other data handling.

    `filepath` should be the complete filepath of the file to be created, including directory and
    suffix.

    `replace_existing`: Flag on whether to overwrite `filepath` if it exists.

    Returns the filepath of the created file.
    """
    filepath = Path(filepath)

    if filepath.is_file() and not replace_existing:
        raise OSError("File `{filepath.name}` already exists and `replace_existing` is false.")

    source_fields, target_fields = set(), set()

    for obj in data:
        source_fields.update(set(obj.get("source", [])))
        target_fields.update(set(obj.get("target", [])))

    source_fields, target_fields = sorted(source_fields), sorted(target_fields)
    source_offset = len(source_fields)

    workbook = xlsxwriter.Workbook(filepath)
    header = workbook.add_format({"bold": True, "font_size": 16})
    orange = workbook.add_format({"bg_color": "#ffdfba", "bold": True, "font_size": 14})
    green = workbook.add_format({"bg_color": "#baffc9", "bold": True, "font_size": 14})
    blue = workbook.add_format({"bg_color": "#bae1ff", "bold": True, "font_size": 14})
    yellow = workbook.add_format({"bg_color": "#ffffba", "bold": True, "font_size": 14})
    red = workbook.add_format({"bg_color": "#ffb3ba", "bold": True, "font_size": 14})
    italic = workbook.add_format({"italic": True, "font_size": 12})
    normal = workbook.add_format({"font_size": 12})

    sheet = workbook.add_worksheet("Matching")

    sheet.write_string(0, 0, "Randonneur template matching file", header)
    sheet.write_string(2, 0, "Metadata", header)
    sheet.write_string(3, 0, "Name", orange)
    sheet.write_string(4, 0, "Description", orange)
    sheet.write_string(4, 2, "String; optional", italic)
    sheet.write_string(5, 0, "License", orange)
    sheet.write_string(6, 0, "Version", orange)
    sheet.write_string(6, 2, 'String like "1.0"; optional', italic)
    sheet.write_string(7, 0, "Source ID", orange)
    sheet.write_string(7, 2, 'String like "SimaPro-9"; optional', italic)
    sheet.write_string(8, 0, "Target ID", orange)
    sheet.write_string(8, 2, 'String like "lcacommons-2024"; optional', italic)
    sheet.write_string(10, 0, "Contributors", yellow)
    sheet.write_string(10, 1, "You can add additional rows via copying if needed", italic)
    sheet.write_string(11, 0, "Name", yellow)
    sheet.write_string(11, 1, "Role", yellow)
    sheet.write_string(11, 2, "Homepage", yellow)
    sheet.write_string(14, 0, "Field mapping", red)
    sheet.write_string(14, 1, "CUSTOM mapping needs to be defined on import", italic)
    sheet.write_string(15, 0, "Source mapping", red)
    sheet.write_string(16, 0, "Target mapping", red)
    sheet.write_string(18, 0, "Data", header)
    template_offset = 19

    sheet.data_validation(5, 1, 5, 1, {"validate": "list", "source": LICENSES})
    sheet.write_string(5, 1, "CC-BY-4.0", normal)
    sheet.data_validation(12, 1, 12, 1, {"validate": "list", "source": ROLES})
    sheet.write_string(12, 1, "author", normal)
    sheet.data_validation(15, 1, 16, 1, {"validate": "list", "source": list(MAPPINGS)})
    sheet.write_string(15, 1, "SIMAPRO_CSV", normal)
    sheet.write_string(16, 1, "ECOSPOLD2", normal)

    sheet.write_string(template_offset, 0, "Source", blue)
    for index in range(1, len(source_fields)):
        sheet.write_string(template_offset, index, "", blue)

    sheet.write_string(template_offset, source_offset, "Target", green)
    for index in range(1, len(target_fields)):
        sheet.write_string(template_offset, index + source_offset, "", green)

    for i, label in enumerate(source_fields):
        sheet.write_string(template_offset + 1, i, label, blue)
    for i, label in enumerate(target_fields):
        sheet.write_string(template_offset + 1, i + source_offset, label, green)
    for row_index, row in enumerate(data):
        for column_index, label in enumerate(source_fields):
            if row["source"].get(label):
                sheet.write_string(
                    template_offset + row_index + 2, column_index, row["source"][label], normal
                )
        for column_index, label in enumerate(target_fields):
            if row["target"].get(label):
                sheet.write_string(
                    template_offset + row_index + 2,
                    source_offset + column_index,
                    row["target"][label],
                    normal,
                )

    sheet.autofit()

    # Shrinks target columns too small
    sheet.set_column(source_offset, source_offset + len(target_fields), 25)

    workbook.close()
    return filepath


def read_excel_template(
    filepath: Path,
    worksheet: str = "Matching",
    license_mapping: Optional[dict] = None,
    field_mapping: Optional[dict] = None,
) -> Datapackage:
    wb = openpyxl.load_workbook(filepath)
    ws = wb[worksheet]

    SECTIONS = {
        "Metadata",
        "Contributors",
        "Field mapping",
        "Data",
    }
    CONTRIBUTORS = {
        "Homepage": "path",
        "Name": "title",
        "Role": "roles",
    }
    METADATA_MAPPING = {
        "Name": "name",
        "Description": "description",
        "License": "licenses",
        "Version": "version",
        "Source ID": "source_id",
        "Target ID": "target_id",
        "Source mapping": "mapping_source",
        "Target mapping": "mapping_target",
    }

    current_section = None
    metadata = {}
    raw_data = defaultdict(list)

    for line in [[cell.value for cell in row] for row in ws.iter_rows(max_row=ws.max_row)]:
        if line[0] == "Randonneur template matching file":
            continue
        elif not any(line):
            continue
        elif line[0] in SECTIONS and current_section != "Data":
            current_section = line[0]
        else:
            raw_data[current_section].append(line)

    for section in SECTIONS:
        if section not in raw_data:
            raise ValueError(f"Missing required section {section}")

    for line in raw_data["Metadata"]:
        key, value = line[0], line[1]
        if key == "License":
            try:
                metadata["licenses"] = [
                    (
                        license_mapping[value]
                        if (isinstance(license_mapping, dict) and value in license_mapping)
                        else ALL_LICENSES[value]
                    )
                ]
            except KeyError as exc:
                raise KeyError(f"Can't find given license short name {value}") from exc
        elif value:
            try:
                metadata[METADATA_MAPPING[key]] = str(value)
            except KeyError as exc:
                raise KeyError(f"Can't understand metadata field {key}") from exc

    for line in raw_data["Field mapping"]:
        if isinstance(field_mapping, dict) and line[1] in field_mapping:
            mapping = field_mapping[line[1]]
        else:
            mapping = getattr(MappingConstants, line[1], None)
            if mapping is None:
                raise KeyError(
                    f"Can't find mapping {line[1]} in built-in or custom field mapping dictionaries"
                )
        try:
            metadata[METADATA_MAPPING[line[0]]] = mapping
        except KeyError as exc:
            raise KeyError(f"Can't understand field mapping field {line[0]}") from exc

    labels = [elem for elem in raw_data["Contributors"][0] if elem]
    metadata["contributors"] = []
    for line in raw_data["Contributors"][1:]:
        try:
            metadata["contributors"].append(
                {
                    CONTRIBUTORS[key]: ([value] if CONTRIBUTORS[key] == "roles" else value)
                    for key, value in zip(labels, line)
                    if value and key
                }
            )
        except KeyError as exc:
            raise KeyError("Can't find given contributor field") from exc

    if "Source" not in raw_data["Data"][0]:
        raise ValueError("Can't find 'Source' in data headers")
    if "Target" not in raw_data["Data"][0]:
        raise ValueError("Can't find 'Target' in data headers")
    source_index, target_index = raw_data["Data"][0].index("Source"), raw_data["Data"][0].index(
        "Target"
    )
    if source_index != 0:
        raise ValueError("'Source' must be the first column in data headers")
    labels = list(
        zip(
            ["source"] * target_index + ["target"] * (len(raw_data["Data"][0]) - target_index),
            raw_data["Data"][1],
        )
    )
    for source_target, label in labels:
        if label not in metadata[f"mapping_{source_target}"]["labels"]:
            given_labels = metadata[f"mapping_{source_target}"]["labels"]
            raise ValueError(
                f"Data label {label} not defined in {source_target} mapping:\n{given_labels}"
            )

    if len(raw_data["Data"]) <= 2:
        raise ValueError("No data found")

    data = []

    for line in raw_data["Data"][2:]:
        this = {"source": {}, "target": {}}
        for (source_target, column), value in zip(labels, line):
            this[source_target][column] = value
        data.append(this)

    rd = Datapackage(**metadata)
    rd.add_data("update", data)
    return rd
